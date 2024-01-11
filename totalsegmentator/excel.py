def set_xlsx_column_width_to_content(file_name):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Load the Excel workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Iterate through the columns and set the width based on the max length of the content (+ 30%)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.3

    # Save the updated workbook
    wb.save(file_name)